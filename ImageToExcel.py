from openpyxl import *
from openpyxl.cell import Cell
from openpyxl.styles import colors
from openpyxl.styles import Color, PatternFill, Font, Border
import PIL
from webcolors import *
from math import *
from tkinter import *
from tkinter import messagebox
from tkinter.ttk import *
from datetime import datetime
def on_button():
    now = datetime.now()
    try:
        fh = open("log.txt", "r")
        log = open("log.txt", "a+")
    except:
        log = open("log.txt", "w+")
        log.write("Start Of The Log History:")
        log.close()
        log = open("log.txt", "a+")
    progressbar.place(x=300, y=215, anchor="center")
    txt.config(state=DISABLED)
    cred.config(state=DISABLED)
    btn.config(state=DISABLED)
    quitter.config(state=DISABLED)
    image = txt.get() + variable.get()
    #print(image)
    alfabeto = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
    wb = Workbook()
    ws1 = wb.active
    #image = input("Image Name: ")
    #print(image)
    file = image + ".xlsx"
    ws1.title = image
    print("")
    try:
        now = datetime.now()
        im = PIL.Image.open(image)
        rgb_im = im.convert('RGB')
        log.write(str(now))
        log.write(": File Name: ")
        log.write(image)
        log.write("\n")
        log.write(str(now))
        log.write(": File Opened Succesfully \n")
    except:
        now = datetime.now()
        messagebox.showerror("Error", "It seems like the image name you put in the field doesn't exist. Try writing it again or make sure that the image is in the same path as this program")
        progressbar.place_forget()
        cred.config(state=NORMAL)
        btn.config(state=NORMAL)
        quitter.config(state=NORMAL)
        txt.config(state=NORMAL)
        log.write(str(now))
        log.write(": Error Opening File \n")
        log.close()
    size = im.size
    print("Width:", size[0])
    print("Height:", size[1])
    print("")
    #print(size)
    nPixel = size[0] * size[1]
    progressbar['maximum'] = nPixel
    #print(nPixel)
    rArr = []
    gArr = []
    bArr = []
    x = 0
    y = 0
    print("Getting colors, this action could take some time depending on the image size")
    for i1 in range (0, size[1]):
        for i2 in range (0, size[0]):
            #print("Coordinate:", x, y)
            r, g, b = rgb_im.getpixel((x, y))
            rArr.append(r)
            gArr.append(g)
            bArr.append(b)
            x = x + 1
        x = 0
        y = y + 1
    now = datetime.now()
    log.write(str(now))
    log.write(": Image's pixel color analized succesfully \n")
    print("")
    print("The image contains", nPixel, "pixels")
    print("")
    index = 0
    index_color = 0
    num = 1
    n_alf = int(len(alfabeto))
    if (size[0] >= len(alfabeto)):
        lettera = floor(size[0] / n_alf)
        l = 0
        for i in range (0, lettera):
            for i in range (0, 26):
                alfabeto.append(alfabeto[l] + alfabeto[i])
            l = l + 1
        #print(alfabeto)
    maxValue = nPixel
    i3 = 0
    print("Styling cells, this action could take some time depending on the image size")
    for i1 in range (0, size[1]):
        index = 0
        for i2 in range (0, size[0]):
            colore = rgb_to_hex((rArr[index_color], gArr[index_color], bArr[index_color]))
            colore_list = list(colore)
            colore_list[0] = "FF"
            colore_argb = "".join(colore_list)
            #print(colore, colore_argb)
            cella = alfabeto[index] + str(num)
            ws1.column_dimensions[alfabeto[index]].width = 3
            c = styles.colors.Color(rgb=colore_argb)
            fillc = styles.fills.PatternFill(patternType='solid', fgColor=c)
            ws1[cella].fill = fillc
            #print("Colorando:", cella, "con colore:", colore)
            index_color = index_color + 1
            index = index + 1
            i3 = i3 + 1
        num = num + 1
        currentValue = nPixel
        progressbar["value"] = (maxValue - (maxValue - i3))
        progressbar.update()
    now = datetime.now()
    log.write(str(now))
    log.write(": Cells Styled Succesfully \n")
    print("")
    print("Process Completed!")
    print("")
    now = datetime.now()
    log.write(str(now))
    log.write(": Process Completed \n")
    messagebox.showinfo("OK!", "Process completed, the file was saved with a 'xlsx' extension")
    wb.save(file)
    progressbar.place_forget()
    cred.config(state=NORMAL)
    btn.config(state=NORMAL)
    quitter.config(state=NORMAL)
    txt.config(state=NORMAL)
    log.close()
    
def func_start(sender):
    on_button()

def cred_start(sender):
    credits()

def credits():
    messagebox.showinfo("About", "Credits: \n Matteo Leggio \n matteo.leggio@tiscali.it")

def f_quitter():
    win.destroy()

def func_quitter(sender):
    f_quitter()
    
def on_closing():
    log = open("log.txt", "a+")
    now = datetime.now()
    log.write(str(now))
    log.write(": Program Closed Prematurely \n")
    log.close()
    win.destroy()

win = Tk()
log = open("log.txt", "a+")
log.write("*******************\n")
log.close()
win.protocol("WM_DELETE_WINDOW", on_closing)
win.resizable(False, False)
win.title("Image To Excel")
win.geometry("600x400")
lbl = Label(win, text="Image To Excel", font=("Verdana", 30, "bold"))
lbl.place(x=300, y=25, anchor="center")
lbl_2 = Label(win, text="Image Name", font=("Verdana", 10, "bold"))
lbl_2.place(x=300, y=155, anchor="center")
desc = Label(win, text="This program transforms an image you give as an input \n into a pixel-art made in Microsoft Excel \n (As long as the image is smaller than 676x676)", font=("Verdana", 10), justify=CENTER)
desc.place(x=300, y=80, anchor="center")
txt = Entry(win ,width=70)
txt.place(x=300, y=175, anchor="center")
variable = StringVar(win)
variable.set(".png")
dropdown = OptionMenu(win, variable, ".png", ".png", ".jpg", ".jpeg")
dropdown.place(x=545, y=175, anchor="center")
progressbar = Progressbar(win,orient="horizontal",length=300,mode="determinate")
btn = Button(win, text="Ok", command=on_button)
btn.place(x=300, y=260, anchor="center")
btn.bind('<Return>', func_start)
cred = Button(win, width=8, text="About", command=credits)
cred.bind('<Return>', cred_start)
cred.place(x=300, y=300, anchor="center")
txt.bind('<Return>', func_start)
quitter = Button(win, width=6, text = "Quit", command=f_quitter)
quitter.place(x=300, y=340, anchor="center")
quitter.bind('<Return>', func_quitter)
win.mainloop()
